"""
task 3
pylint 10
"""
from openpyxl import Workbook

# Создаем первый Excel файл (1111.xlsx) и записываем числа
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Sheet1"
data1 = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
for row in data1:
    ws1.append(row)
wb1.save("1111.xlsx")

# Создаем второй Excel файл (2222.xlsx) и записываем числа
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Sheet1"
data2 = [
    [9, 8, 7],
    [6, 5, 4],
    [3, 2, 1]
]
for row in data2:
    ws2.append(row)
wb2.save("2222.xlsx")

# Создаем третий Excel файл (3333.xlsx) и записываем числа
wb3 = Workbook()
ws3 = wb3.active
ws3.title = "Sheet1"
data3 = [
    [5, 1, 8],
    [2, 9, 4],
    [7, 3, 6]
]
for row in data3:
    ws3.append(row)
wb3.save("3333.xlsx")

print("Три Excel файла успешно созданы и заполнены числами.")
